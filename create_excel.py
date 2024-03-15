# Author: Cesium06

# Example of chopping up an excel with an irregular format. Headers scattered throughout excel



from typing import TextIO

import xlrd
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import os
from pathlib import Path


def is_workbook_open_for_writing(workbook_path):
    if os.path.exists(workbook_path):
        try:
            # Try to open the workbook in write mode
            f: TextIO
            with open(workbook_path, 'w') as f:
                pass
            return False  # Workbook is not open for writing
        except PermissionError:
            return True  # Workbook is open for writing
    else:
        return False  # File does not exist


class Create_Excel:
    input_excel_file = ''
    output_excel = ''
    workbook = ''
    set_of_users = set()
    target_value = 'Issue Date'
    row_data_array = dict()
    sheet = ''

    def __init__(self, input_excel_file):
        self.input_excel_file = input_excel_file
        self.workbook = xlrd.open_workbook(input_excel_file)
        self.sheet = self.workbook.sheet_by_index(0)

    def get_excel_name(self):
        row_index = 4  # Rows are zero-indexed, so the 5th row has index 4
        row_values = self.sheet.row_values(row_index)

        # Assuming the date range is in the first cell of the 5th row
        date_range_str = row_values[0]

        # Extract the start and end dates from the string
        date_range_parts = date_range_str.split(' through ')
        start_date_str = date_range_parts[0].split(' ')[-1]  # Extract the date part after the last space
        end_date_str = date_range_parts[1].split(' ')[-1]  # Extract the date part after the last space

        # Convert date strings to datetime objects
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date().strftime('%m%d%Y')
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date().strftime('%m%d%Y')

        self.output_excel = f"TS_{start_date}_{end_date}.xlsx"

    def generate_excel_and_workbooks(self):
        # Initialize a list to store the integer values from the rows preceding each occurrence of the target value
        preceding_int_values = []
        integer_value = 0

        # Iterate over rows
        for row_index in range(self.sheet.nrows):
            # Initialize a flag to skip empty rows after finding preceding_int_values
            skip_empty_row = False

            # Iterate over cells in the row
            for col_index in range(self.sheet.ncols):
                # Get the value of the current cell
                cell_value = self.sheet.cell_value(row_index, col_index)

                # Check if the cell value matches the target value
                if cell_value == self.target_value:
                    # Determine the index of the row preceding the row where the target value is found
                    preceding_row_index = row_index - 1
                    # If the preceding row index is valid, retrieve the value from the specified column
                    if preceding_row_index >= 0:
                        preceding_value = self.sheet.cell_value(preceding_row_index,
                                                                col_index)  # Assuming same column as target value
                        # Extract integer value using regular expression
                        integer_value = re.search(r'\d+', str(preceding_value))
                        if integer_value:
                            key = int(integer_value.group())
                            if key not in self.row_data_array:
                                self.row_data_array[key] = []
                        else:
                            # Handle the case where no integer value is found
                            print(row_index)
                    else:
                        # Handle the case where the target value is found in the first row
                        preceding_int_values.append(None)

                    # Set the flag to skip empty rows
                    skip_empty_row = True
                    # Break the inner loop to avoid checking other cells in the same row
                    break

            # If skip_empty_row flag is set and the row is empty, skip it
            if skip_empty_row and all(
                    self.sheet.cell_value(row_index, col_index) == '' for col_index in range(self.sheet.ncols)):
                continue

            # Fetch the row data here and process it as needed
            row_data = [self.sheet.cell_value(row_index, col_index) for col_index in range(self.sheet.ncols)]
            filtered_list = [item for item in row_data if item != '']
            if filtered_list:
                if integer_value:
                    key = int(integer_value.group())
                    self.set_of_users.add(key)
                    self.row_data_array[key].append(tuple(filtered_list))

        set_of_users_sorted = sorted(self.set_of_users)
        set_of_users_sorted.append('Total')

        # Create a new workbook
        wb = Workbook()

        # Create a new sheet for each set of integer values
        for i, value in enumerate(set_of_users_sorted):
            if value is not None:
                sheet_name = f"{value}"
                ws = wb.create_sheet(title=sheet_name)
                # Write the value to the first cell of the sheet
                ws.cell(row=1, column=1, value="EO")
                ws.cell(row=1, column=2, value="Issue Date")
                ws.cell(row=1, column=3, value="Tickets #")
                ws.cell(row=1, column=4, value="State")
                ws.cell(row=1, column=5, value="License")
                ws.cell(row=1, column=6, value="Location")
                ws.cell(row=1, column=7, value="Violation")
                ws.cell(row=1, column=8, value="Warning")
                ws.cell(row=1, column=9, value="Void")
            if 'Sheet' in wb:
                wb.remove(wb['Sheet'])

        # Save the workbook to a new Excel file
        wb.save('Output/' + self.output_excel)

        print(f"Excel sheets created successfully. File saved as Output/{self.output_excel}")

    def populate_excel(self):
        global sheet_to_be_active_in
        sheet_location = {}
        date_time_regex = r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}"
        df = pd.read_excel(self.input_excel_file)
        wb = load_workbook(Path('Output/' + self.output_excel))
        current_sheet = ''

        for index, row in df.iterrows():
            if re.match(date_time_regex, str(row['Unnamed: 1'])):
                # Select the worksheet you want to write to
                ws = wb[str(sheet_to_be_active_in)]

                # Now we write each row to the correct sheet!
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=1, value=sheet_to_be_active_in)
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=2, value=row['Unnamed: 1'])
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=3, value=row['Unnamed: 4'])
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=4, value=row['Unnamed: 8'])
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=5, value=row['Unnamed: 10'])
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=6, value=row['Unnamed: 13'])
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=7, value=row['Unnamed: 19'])
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=8, value=row['Unnamed: 22'])
                ws.cell(row=sheet_location[sheet_to_be_active_in], column=9, value=row['Unnamed: 23'])

                sheet_location[sheet_to_be_active_in] += 1
            else:
                # Define the regex pattern
                user_pattern = r"\b(\d+)"

                # Match the pattern
                match = re.match(user_pattern, str(row['Unnamed: 1']).strip())

                # Extract the value if there's a match
                if match:
                    sheet_to_be_active_in = match.group(1)

                    if sheet_to_be_active_in != current_sheet or current_sheet == '':
                        if sheet_to_be_active_in in sheet_location:
                            sheet_location[sheet_to_be_active_in] = sheet_location[sheet_to_be_active_in] + 1
                            current_sheet = sheet_to_be_active_in
                        else:
                            sheet_location[sheet_to_be_active_in] = 2
                            current_sheet = sheet_to_be_active_in

        # Save the workbook to a new Excel file (outside the loop)
        wb.save('Output/' + self.output_excel)


        # Loop once more to populate total

        # Select the worksheet you want to write to
        ws = wb[str('Total')]
        df = pd.read_excel(self.input_excel_file)
        internal_index = 2
        for index, row in df.iterrows():
            if re.match(date_time_regex, str(row['Unnamed: 1'])):
                # Now we write each row to the correct sheet!
                ws.cell(row=internal_index, column=1, value=sheet_to_be_active_in)
                ws.cell(row=internal_index, column=2, value=row['Unnamed: 1'])
                ws.cell(row=internal_index, column=3, value=row['Unnamed: 4'])
                ws.cell(row=internal_index, column=4, value=row['Unnamed: 8'])
                ws.cell(row=internal_index, column=5, value=row['Unnamed: 10'])
                ws.cell(row=internal_index, column=6, value=row['Unnamed: 13'])
                ws.cell(row=internal_index, column=7, value=row['Unnamed: 19'])
                ws.cell(row=internal_index, column=8, value=row['Unnamed: 22'])
                ws.cell(row=internal_index, column=9, value=row['Unnamed: 23'])
                internal_index += 1

            else:
                # Define the regex pattern
                user_pattern = r"\b(\d+)"

                # Match the pattern
                match = re.match(user_pattern, str(row['Unnamed: 1']).strip())

                # Extract the value if there's a match
                if match:
                    sheet_to_be_active_in = match.group(1)

                    # if sheet_to_be_active_in != current_sheet or current_sheet == '':
                    #     if sheet_to_be_active_in in sheet_location:
                    #         sheet_location[sheet_to_be_active_in] = sheet_location[sheet_to_be_active_in] + 1
                    #         current_sheet = sheet_to_be_active_in
                    #     else:
                    #         sheet_location[sheet_to_be_active_in] = 2
                    #         current_sheet = sheet_to_be_active_in

        # Save the workbook to a new Excel file (outside the loop)
        wb.save('Output/' + self.output_excel)


